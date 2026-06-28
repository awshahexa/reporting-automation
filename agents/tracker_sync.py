"""Tracker sync — read WBS tracker Excel files and update DB."""
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from agents.database import DatabaseAgent


def detect_node_type(filename):
    """Detect IPCORE/TRUNK/ACC/RNC from filename."""
    name = filename.upper()
    if "IPCORE" in name:
        return "IPCORE"
    if "TRUNK" in name:
        return "TRUNK"
    if "RNC" in name:
        return "RNC"
    if "ACC" in name:
        return "ACC"
    return None


def sync_tracker_file(filepath, db=None):
    """Sync a WBS tracker Excel file into site_activities and derive milestones."""
    filepath = Path(filepath)
    fname = filepath.name
    node_type = detect_node_type(fname)
    if not node_type:
        return {"status": "error", "message": f"Cannot detect node type from filename: {fname}"}

    print(f"\nSyncing {node_type} tracker: {fname}")
    db = db or DatabaseAgent()
    conn = db.connect()

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb["data"]

    rows_iter = iter(ws.iter_rows(values_only=False))
    headers = [c.value for c in next(rows_iter)]
    sub_headers = [c.value for c in next(rows_iter)]
    field_types = [c.value for c in next(rows_iter)]
    field_names_raw = [str(c.value or "").strip() for c in next(rows_iter)]

    # Site code is in column 0
    # Build readable field names from sub-headers + field types + field names
    # Format: sub > [ftype >] fn (ftype included when it disambiguates columns)
    field_names = {}
    col_keys = []
    for ci, (sub, ftype, fname) in enumerate(zip(sub_headers, field_types, field_names_raw)):
        if ci == 0:
            field_names[ci] = "site_code"
        else:
            sub_str = str(sub or "").strip()
            ftype_str = str(ftype or "").strip()
            fn_str = str(fname or "").strip()
            if sub_str and fn_str:
                if ftype_str and ftype_str not in (sub_str, fn_str) and fn_str not in ftype_str:
                    field_names[ci] = f"{sub_str} > {ftype_str} > {fn_str}"
                else:
                    field_names[ci] = f"{sub_str} > {fn_str}"
            elif sub_str and ftype_str and ftype_str not in sub_str:
                field_names[ci] = f"{sub_str} > {ftype_str}"
            elif fn_str:
                field_names[ci] = fn_str
            elif ftype_str:
                field_names[ci] = ftype_str
            else:
                field_names[ci] = f"col_{ci}"
        col_keys.append(field_names[ci])

    # Existing data for change detection
    existing = {}
    for row in conn.execute(
        "SELECT site_code, node_type, activity_name, field_name, value FROM site_activities WHERE node_type=?",
        (node_type,)
    ).fetchall():
        existing[(row["site_code"], row["activity_name"], row["field_name"])] = row["value"]

    site_node_types = {}
    for row in conn.execute("SELECT site_code, node_type FROM sites").fetchall():
        nt = (row["node_type"] or "").strip()
        if nt:
            site_node_types[row["site_code"]] = set(filter(None, nt.split(",")))

    # Pre-read all rows (extract cell values for data)
    data_rows = [[c.value for c in row] for row in rows_iter]

    changes = 0
    sites_seen = set()

    for row in data_rows:
        sc = str(row[0] or "").strip()
        if not sc:
            continue
        sc = sc.upper()
        sites_seen.add(sc)

        # Update site node_type
        current_types = site_node_types.get(sc, set())
        current_types.add(node_type)
        site_node_types[sc] = current_types
        conn.execute("""
            INSERT INTO sites (site_code, site_name, node_type, updated_at)
            VALUES (?, ?, ?, datetime('now'))
            ON CONFLICT(site_code) DO UPDATE SET
                node_type=excluded.node_type, updated_at=datetime('now')
        """, (sc, sc, ",".join(sorted(current_types))))

        for ci in range(1, len(row)):
            val = str(row[ci] or "").strip()
            fname = field_names.get(ci, f"col_{ci}")
            act_name, _, field_name = fname.partition(" > ")
            if not field_name:
                field_name = act_name
                act_name = "IP Bearer Network"

            old_val = existing.get((sc, act_name, field_name))
            if old_val != val:
                if old_val is not None:
                    conn.execute("""
                        INSERT INTO tracker_audit (site_code, node_type, activity_name, field_name, old_value, new_value, changed_at)
                        VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                    """, (sc, node_type, act_name, field_name, str(old_val), str(val)))
                conn.execute("""
                    INSERT INTO site_activities (site_code, node_type, activity_name, field_name, value, updated_at)
                    VALUES (?, ?, ?, ?, ?, datetime('now'))
                    ON CONFLICT(site_code, node_type, activity_name, field_name) DO UPDATE SET
                        value=excluded.value, updated_at=datetime('now')
                """, (sc, node_type, act_name, field_name, val))
                existing[(sc, act_name, field_name)] = val
                changes += 1

    # ── Milestone Derivation (based on actual dates in site_activities) ──
    print(f"  Deriving milestones for {node_type}...")
    ms_derived = 0

    for sc in sites_seen:
        # --- Three-state milestone derivation ---
        # CC: Not started -> Ongoing (TSSR approved) -> Completed (Site PAC Approval actual date)
        tssr_done = conn.execute("""
            SELECT 1 FROM site_activities
            WHERE site_code=? AND activity_name='Technical Site Survey'
              AND field_name IN ('TSS > actual end time','eTSS SCF Approval > actual end time','eTSS ZTE Approval > actual end time')
              AND value IS NOT NULL AND value != ''
            LIMIT 1
        """, (sc,)).fetchone()
        cc_done = conn.execute("""
            SELECT 1 FROM site_activities
            WHERE site_code=? AND activity_name='Commisioning Certificate'
              AND field_name='Site PAC Approval > actual end time'
              AND value IS NOT NULL AND value != ''
            LIMIT 1
        """, (sc,)).fetchone()
        if cc_done:
            cc_status = "Completed"
        elif tssr_done:
            cc_status = "Ongoing"
        else:
            cc_status = "Not started"

        # PAC: Not started -> Ongoing (CC completed) -> Completed (PAC Approval actual date)
        pac_done = conn.execute("""
            SELECT 1 FROM site_activities
            WHERE site_code=? AND activity_name='(Preliminary/Provisional) Acceptance Certification'
              AND field_name='PAC Approval > actual end time'
              AND value IS NOT NULL AND value != ''
            LIMIT 1
        """, (sc,)).fetchone()
        if pac_done:
            pac_status = "Completed"
        elif cc_done:
            pac_status = "Ongoing"
        else:
            pac_status = "Not started"

        # FAC: Not started -> Ongoing (PAC completed) -> Completed (FAC Approval actual date)
        fac_done = conn.execute("""
            SELECT 1 FROM site_activities
            WHERE site_code=? AND activity_name='Final Acceptance Certification'
              AND field_name='FAC Approval > actual end time'
              AND value IS NOT NULL AND value != ''
            LIMIT 1
        """, (sc,)).fetchone()
        if fac_done:
            fac_status = "Completed"
        elif pac_done:
            fac_status = "Ongoing"
        else:
            fac_status = "Not started"

        conn.execute("""
            INSERT INTO milestone_status (site_code, milestone, status, updated_at)
            VALUES (?, 'CC', ?, datetime('now'))
            ON CONFLICT(site_code, milestone) DO UPDATE SET
                status=excluded.status, updated_at=datetime('now')
        """, (sc, cc_status))
        ms_derived += 1

        conn.execute("""
            INSERT INTO milestone_status (site_code, milestone, status, updated_at)
            VALUES (?, 'PAC', ?, datetime('now'))
            ON CONFLICT(site_code, milestone) DO UPDATE SET
                status=excluded.status, updated_at=datetime('now')
        """, (sc, pac_status))
        ms_derived += 1

        conn.execute("""
            INSERT INTO milestone_status (site_code, milestone, status, updated_at)
            VALUES (?, 'FAC', ?, datetime('now'))
            ON CONFLICT(site_code, milestone) DO UPDATE SET
                status=excluded.status, updated_at=datetime('now')
        """, (sc, fac_status))
        ms_derived += 1

    conn.commit()
    conn.close()
    wb.close()

    return {
        "status": "ok",
        "node_type": node_type,
        "sites": len(sites_seen),
        "changes": changes,
        "milestones_derived": ms_derived,
    }
