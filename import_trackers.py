"""Import all 3 tracker Excel files into the reporting database, building
the site database grouped by node type (IPCORE, TRUNK, ACC)."""
import sys, sqlite3
from pathlib import Path
from agents.database import DatabaseAgent
from agents.config import MILESTONES

HOT_FOLDER = Path(__file__).parent / "sharepoint_sim" / "Working" / "Sites Document" / "Submit (hot folder)"

TRACKER_FILES = {}
for f in sorted(HOT_FOLDER.glob("*.xlsx")):
    name = f.name.upper()
    if "IPCORE" in name:
        TRACKER_FILES["IPCORE"] = str(f)
    elif "TRUNK" in name:
        TRACKER_FILES["TRUNK"] = str(f)
    elif "RNC" in name:
        TRACKER_FILES["RNC"] = str(f)
    elif "ACC" in name:
        TRACKER_FILES["ACC"] = str(f)


def parse_tracker(filepath):
    """Parse one tracker Excel file and return list of site dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["data"]

    # Read rows - iter_rows with explicit range
    max_c = min(ws.max_column, 80)
    raw = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_c, values_only=True):
        raw.append([str(c or "").strip() for c in row])

    # Find the header row (contains "customer site code")
    header_idx = None
    for i, row in enumerate(raw):
        if any("customer site code" in v.lower() for v in row):
            header_idx = i
            break
    if header_idx is None:
        print("  WARNING: Could not find header row with 'customer site code'")
        return []

    for i in range(header_idx + 1):
        vals = [v for v in raw[i] if v]
        if vals:
            print(f"  raw[{i}]: {vals[:8]}")

    headers = raw[header_idx]

    # Identify columns from headers
    site_code_col = site_name_col = region_col = None
    du_code_col = du_name_col = cluster_col = None

    for c, h in enumerate(headers):
        h_lower = h.lower()
        if h_lower == "customer site code":
            site_code_col = c
        elif h_lower == "customer site name":
            site_name_col = c
        elif h_lower == "region":
            region_col = c
        elif h_lower == "du code":
            du_code_col = c
        elif h_lower == "du name":
            du_name_col = c
        elif h_lower == "cluster":
            cluster_col = c

    # Scan category rows (before header) to identify activity status columns
    status_cols = {}
    if header_idx >= 2:
        cat_row = raw[header_idx - 2]  # row with category labels
        sub_row = raw[header_idx - 1]  # row with sub-category labels
        for c in range(len(headers)):
            cat = cat_row[c] if c < len(cat_row) else ""
            sub = sub_row[c] if c < len(sub_row) else ""
            hdr = headers[c]
            if cat and cat != "site basic info" and sub == cat and hdr == "status":
                status_cols[cat] = c

    if site_code_col is None:
        print(f"  WARNING: Could not find site_code column in {filepath}")
        return []

    sites = []
    for r in range(header_idx + 1, len(raw)):
        row = raw[r]
        code = row[site_code_col] if site_code_col < len(row) else ""
        if not code:
            continue

        site = {
            "site_code": code,
            "site_name": row[site_name_col] if site_name_col and site_name_col < len(row) else code,
            "region": row[region_col] if region_col and region_col < len(row) else "",
            "du_code": row[du_code_col] if du_code_col and du_code_col < len(row) else "",
            "du_name": row[du_name_col] if du_name_col and du_name_col < len(row) else "",
            "cluster": row[cluster_col] if cluster_col and cluster_col < len(row) else "",
        }

        # Capture activity statuses for progress tracking
        for act, col in status_cols.items():
            if col < len(row):
                site[f"act_{act}"] = row[col]

        sites.append(site)

    wb.close()
    return sites


def import_all_trackers():
    db = DatabaseAgent()

    all_sites = {}  # site_code -> merged site data
    node_map = {}   # site_code -> set of node types

    for node_type, filepath in TRACKER_FILES.items():
        fname = Path(filepath).name.encode('utf-8', errors='replace').decode('utf-8')
        print(f"\nParsing {node_type}: {fname}")
        sites = parse_tracker(filepath)
        print(f"  Found {len(sites)} sites")

        for s in sites:
            sc = s["site_code"]
            if sc not in all_sites:
                all_sites[sc] = s
                node_map[sc] = set()
            else:
                # Merge: prefer non-empty values
                for k in ["site_name", "region", "du_code", "du_name", "cluster"]:
                    if s.get(k) and not all_sites[sc].get(k):
                        all_sites[sc][k] = s[k]
            node_map[sc].add(node_type)

    print(f"\n{'='*60}")
    print(f"Total unique sites across all trackers: {len(all_sites)}")

    # Batch upsert into database (single connection for speed)
    db = DatabaseAgent()
    conn = db.connect()
    conn.execute("BEGIN")
    upserted = 0
    for sc, s in all_sites.items():
        node_types = ",".join(sorted(node_map[sc]))
        conn.execute("""
            INSERT INTO sites (site_code, site_name, region, du_code, du_name, node_type)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(site_code) DO UPDATE SET
                site_name=excluded.site_name, region=excluded.region,
                du_code=excluded.du_code, du_name=excluded.du_name,
                node_type=excluded.node_type, updated_at=datetime('now')
        """, (sc, s.get("site_name", sc), s.get("region"), s.get("du_code"), s.get("du_name"), node_types))

        for ms in MILESTONES:
            conn.execute("""
                INSERT INTO milestone_status (site_code, milestone, status, updated_by, updated_at)
                VALUES (?, ?, 'Not started', 'tracker_import', datetime('now'))
                ON CONFLICT(site_code, milestone) DO UPDATE SET
                    status='Not started', updated_by='tracker_import', updated_at=datetime('now')
            """, (sc, ms))

        upserted += 1

    conn.commit()
    conn.close()
    print(f"Upserted {upserted} sites into database")

    # Log node type distribution
    from collections import Counter
    node_counts = Counter()
    for sc, nset in node_map.items():
        node_counts[",".join(sorted(nset))] += 1
    print(f"\nNode type distribution:")
    for combo, cnt in node_counts.most_common():
        print(f"  {combo}: {cnt} sites")

    # Show totals per node type
    print(f"\nPer node type:")
    for nt in ["IPCORE", "TRUNK", "ACC"]:
        count = sum(1 for nset in node_map.values() if nt in nset)
        print(f"  {nt}: {count} sites")

    return all_sites, node_map


def import_from_tracker_db():
    """Fallback: if we already have a tracker database from the dashboard,
    migrate from there."""
    tracker_db = Path(__file__).parent.parent / "Tracker"
    # No separate DB exists - trackers are Excel files only
    pass


if __name__ == "__main__":
    print("=" * 60)
    print("  SACOFA Reporting Automation - Tracker Import")
    print("=" * 60)

    all_sites, node_map = import_all_trackers()

    db = DatabaseAgent()
    stats = db.get_summary_stats()
    print(f"\n{'='*60}")
    print(f"Database summary:")
    print(f"  Total sites: {stats['total_sites']}")
    print(f"  Total documents: {stats['total_documents']}")
    print(f"\nDone. Run 'python run.py start' to launch the dashboard.")
