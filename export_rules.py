"""Export validation rules matrix to Excel (v5)."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
default_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
specific_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_row(ws, row, cols, fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = wrap
        cell.border = thin_border
        if fill:
            cell.fill = fill


# ═══════════════════════════════════════════
# Sheet 1: Rules Summary
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "Rules Summary"

ws.merge_cells("A1:D1")
ws.cell(row=1, column=1, value="Default Rules (applied to ALL document types)").font = Font(bold=True, size=13, color="FFFFFF")
ws.cell(row=1, column=1).fill = header_fill
ws.cell(row=1, column=1).alignment = header_align

headers = ["#", "Rule Name", "Description", "Pass / Fail Criteria"]
for i, h in enumerate(headers, 1):
    ws.cell(row=2, column=i, value=h)
style_header(ws, 2, 4)

data = [
    [1, "Valid PDF", "_rule_pdf_valid", "File must be readable by PyMuPDF (fitz) and have >= 1 page"],
    [2, "Has Text", "_rule_has_text", "Extractable text >= 10 chars; OCR fallback (pytesseract) if extracted < 50 chars"],
    [3, "File Naming", "_rule_file_naming", "Must match [DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf pattern"],
]
for r, row in enumerate(data, 3):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_row(ws, r, 4, default_fill)

# Doc-Type-Specific Rules
start = 8
ws.merge_cells(f"A{start}:H{start}")
ws.cell(row=start, column=1, value="Doc-Type-Specific Rules").font = Font(bold=True, size=13, color="FFFFFF")
ws.cell(row=start, column=1).fill = header_fill
ws.cell(row=start, column=1).alignment = header_align

headers2 = ["Doc Type", "Milestone", "Rules", "Content Details", "Blur / Photo / Sign Rules", "Output Destination (Review)", "File Naming Pattern"]
for i, h in enumerate(headers2, 1):
    ws.cell(row=start+1, column=i, value=h)
style_header(ws, start+1, 7)

specific = [
    # ── CC milestone ──
    ["TSSR", "CC",
     "_rule_tssr_content\n_rule_no_blurry_pages\n_rule_min_photos_per_category\n_rule_has_signature",
     "7 fields (SITE NAME, SITE OWNER, TOWER ID, REGION, SITE STRUCTURE, EQUIPMENT MODEL, EQUIPMENT FUNCTION); Lat/Lng label; NMS LRD sitename match filename; OCR fallback",
     "Blur: Lap var >=300 @100dpi, resize 500px, skip blank\nPhotos: per-section heading, >=1 photo (200x200px), exclude page 1\nSign: full doc, N/A exclusion",
     "Review/[SITE]/CC/",
     "[SITENAME]_TSSR_[NODETYPE]_[YYYYMMDD].pdf"],
    ["SATP", "CC",
     "_rule_satp_content\n_rule_no_blurry_pages\n_rule_has_signature",
     "Full OCR (scanned doc); 4 fields (SITE NAME, SITE OWNER, NMS LRD, LAT/LONG alt. LATITUDE+LONGITUDE); signing page (SATP INFO + Prepared by + Verified by + Date); date not future; sitename match (colon or space)",
     "Blur: Lap var >=300 @100dpi, resize 500px, skip blank\nSign: OCR text fallback for 'Signature'/'Signed'",
     "Review/[SITE]/CC/",
     "[SITENAME]_SATP_[NODETYPE]_[YYYYMMDD].pdf"],
    ["PO", "CC, PAC, FAC",
     "_rule_po_content\n_rule_has_signature (last page)",
     "Purchase Order heading; PO number; Vendor/Order To; Date (dd/mm/yyyy); site code in doc; OCR fallback",
     "Sign: last page only; embedded image detection + OCR N/A placeholder exclusion",
     "Review/[SITE]/CC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["DN", "CC",
     "_rule_dn_content\n_rule_has_signature (last page)",
     "DN/DO number; Date (dd/mm/yyyy or yyyymmdd); Customer (SacoFA / IX Telecom); site code match; Consignee/Warehouse Manager signature zone; OCR fallback",
     "Sign: last page only; embedded image detection + OCR N/A placeholder exclusion",
     "Review/[SITE]/CC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["BL", "CC",
     "_rule_bl_content",
     "Container/BL number; Shipper (ZTE CORPORATION); Consignee (ZTE Malaysia); Port of Loading/Discharge; Vessel name; OCR fallback",
     "\u2014",
     "Review/[SITE]/CC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["PL", "CC (supporting)",
     "_rule_pl_content\n_rule_has_signature (last page)",
     "Reference (PR/SOR/YYYY/NNNN); PO No (SSB/PO/NNNN-NN); PO Date (opt.); Consignee; Project; site match; Prepared By; Date Prepared (opt.); Total Quantity; OCR fallback",
     "Sign: last page only; embedded image detection + N/A placeholder exclusion via OCR",
     "Review/[SITE]/CC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["FT", "CC (supporting)",
     "_rule_ft_content\n_rule_no_blurry_pages\n_rule_has_signature (start_page=6)",
     "Functionality Test heading; Node Name containing site code; date; OCR fallback",
     "Blur: Lap var >=1000 @100dpi (default threshold)\nSign: scan from page 6 onward (first 5 are cover/summary); N/A placeholder exclusion via OCR",
     "Review/[SITE]/CC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    # ── PAC milestone ──
    ["CC (Cert)", "PAC",
     "_rule_cc_content\n_rule_has_signature (last page)",
     "Commissioning/Completion Certificate heading; NMS LRD site match; acceptance/signatory zone; OCR fallback",
     "Sign: last page only; embedded image detection + N/A exclusion",
     "Review/[SITE]/PAC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["MOP", "PAC",
     "_rule_mop_content\n_rule_has_signature (full doc)",
     "Method of Procedure heading; Version/Revision; date; site code; signatory label (Prepared/Verified/Approved By); OCR fallback",
     "Sign: full doc; embedded image detection + N/A placeholder exclusion via OCR",
     "Review/[SITE]/PAC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["SMR", "PAC",
     "_rule_smr_content\n_rule_has_signature (page 2 only)",
     "Service Migration Report heading; NMS LRD site match; 8 page-1 fields (SITE NAME, SITE OWNER, LATITUDE, LONGITUDE, REGION, SITE STRUCTURE, EQUIPMENT MODEL, EQUIPMENT FUNCTION); 3 signatory blocks (Prepared/Verified/Approved By); OCR fallback",
     "Sign: page 2 only; embedded image detection + N/A placeholder exclusion via OCR",
     "Review/[SITE]/PAC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    ["PAF", "PAC, FAC",
     "_rule_paf_content\n_rule_has_signature (last page)",
     "Provisional Acceptance Certificate heading; NMS LRD site match; acceptance/signatory zone; OCR fallback",
     "Sign: last page only; embedded image detection + N/A placeholder exclusion via OCR",
     "Review/[SITE]/PAC/, Review/[SITE]/FAC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
    # ── FAC milestone ──
    ["FAF", "FAC",
     "_rule_faf_content\n_rule_has_signature (last page)",
     "Final Acceptance Certificate heading; NMS LRD site match; signatory/acceptance zone; OCR fallback",
     "Sign: last page only; embedded image detection + N/A placeholder exclusion via OCR",
     "Review/[SITE]/FAC/",
     "[DOCTYPE]_[SITENAME]_[NODETYPE]_[YYYYMMDD].pdf"],
]

for r, row in enumerate(specific, start+2):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_row(ws, r, 7, specific_fill)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 42
ws.column_dimensions["E"].width = 38
ws.column_dimensions["F"].width = 28
ws.column_dimensions["G"].width = 42

# ═══════════════════════════════════════════
# Sheet 2: Rule Definitions
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Rule Definitions")

ws2.merge_cells("A1:D1")
ws2.cell(row=1, column=1, value="Rule Definitions - Full Descriptions").font = Font(bold=True, size=13, color="FFFFFF")
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=1).alignment = header_align

h2 = ["Rule Name", "Line in File", "Purpose", "Implementation Details"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header(ws2, 2, 4)

rules_def = [
    ["_rule_pdf_valid", "sharepoint_manager.py:536",
     "Check file is a readable PDF",
     "Opens with fitz.open(); returns False if exception or page_count == 0"],
    ["_rule_has_text", "sharepoint_manager.py:550",
     "Check PDF contains extractable text; fallback to OCR for scanned docs",
     "Extracts text from all pages via get_text('text'); if < 50 chars, runs pytesseract OCR per page at 200dpi; final check: len(text.strip()) >= 10"],
    ["_rule_file_naming", "sharepoint_manager.py:783",
     "Validate filename convention",
     "Regex: [A-Za-z0-9\\s]+_[A-Za-z0-9]+_[A-Za-z\\s]+_\\d{8}\\.pdf"],
    ["_rule_tssr_content", "sharepoint_manager.py:792",
     "Validate TSSR page 1 fields and sitename",
     "Parse filename (new sitename-first or legacy); extract 7 required fields by label-position (look-ahead 5 lines); check Latitude/Longitude label; extract NMS LRD/Contract No value -> compare to filename sitename; OCR fallback if page 1 text < 50 chars"],
    ["_rule_satp_content", "sharepoint_manager.py:929",
     "Validate SATP scanned form content",
     "Full OCR all pages at 200dpi; page 1: check 4 required fields (SITE NAME, SITE OWNER, NMS LRD, LAT/LONG with LATITUDE+LONGITUDE alternate); signing page: find SATP INFO + Prepared by/Surveyor name + Verified by + Date; validate sign-off date not in future; sitename from NMS LRD (colon or space-separated) matches filename"],
    ["_rule_pl_content", "sharepoint_manager.py:1072",
     "Validate PL/Packing List content",
     "OCR-tolerant regex: Reference (PR/SOR/YYYY/NNNN or similar), PO No (SSB/PO/NNNN-NN), PO Date (optional), site match (same-line, next-line, or full-text fallback), Prepared By text block, Date Prepared (optional), Consignee (e.g. SACOFA), Project name, Total Quantity; OCR fallback when text < 50 chars"],
    ["_rule_po_content", "sharepoint_manager.py:1313",
     "Validate PO content; signature rule separate (last page)",
     "Purchase Order heading or PO number pattern; Vendor via 'Order To' label or VENDOR/SUPPLIER; date via dd/mm/yyyy regex; site code match in document text; OCR fallback for scanned POs"],
    ["_rule_dn_content", "sharepoint_manager.py:1420",
     "Validate DN/Delivery Note content",
     "DN/DO # or Delivery Note heading; date (dd/mm/yyyy, yyyymmdd, or month-name); customer via SACOFA/IX Telecom/Consignee; site code (soft match for OCR-merged text); Consignee/Warehouse Manager signature zone; OCR fallback"],
    ["_rule_bl_content", "sharepoint_manager.py:1592",
     "Validate Bill of Lading content",
     "Container/BL number pattern; Shipper (ZTE CORPORATION); Consignee (ZTE Malaysia); Port of Loading/Discharge (by name or label); Vessel name; OCR fallback"],
    ["_rule_mop_content", "sharepoint_manager.py:1584",
     "Validate MOP content",
     "Method of Procedure heading; Version/Revision number; date; site code match; signatory label (Prepared/Verified/Approved By); OCR fallback"],
    ["_rule_smr_content", "sharepoint_manager.py:1658",
     "Validate SMR (Service Migration Report) content",
     "Service Migration Report heading; NMS LRD value matches filename sitename; 8 required fields on page 1 (SITE NAME, SITE OWNER, LATITUDE, LONGITUDE, REGION, SITE STRUCTURE, EQUIPMENT MODEL, EQUIPMENT FUNCTION); 3 signatory blocks (Prepared/Verified/Approved By); OCR fallback"],
    ["_rule_cc_content", "sharepoint_manager.py:1782",
     "Validate CC (Commissioning/Completion Certificate) content",
     "Commissioning or Completion Certificate heading; NMS LRD site match; signatory/acceptance zone (Certified/Approved By or Acceptance Date/Remarks); OCR fallback"],
    ["_rule_ft_content", "sharepoint_manager.py:1984",
     "Validate FT (Functionality Test) content; blur detection added separately",
     "Functionality Test heading; Node Name containing site code; date; signatory label (including Tested By); OCR fallback"],
    ["_rule_paf_content", "sharepoint_manager.py:2084",
     "Validate PAF (PAC Approval Form) content",
     "Provisional Acceptance Certificate heading; NMS LRD site match; signatory/acceptance zone; OCR fallback"],
    ["_rule_faf_content", "sharepoint_manager.py:2048",
     "Validate FAF (FAC Acceptance Form) content",
     "Final Acceptance Certificate heading; NMS LRD site match; signatory/acceptance zone; OCR fallback"],
    ["_rule_no_blurry_pages", "sharepoint_manager.py:654",
     "Detect blurry pages via Laplacian variance",
     "Renders each page at 100dpi -> OpenCV grayscale -> optional resize (SATP/TSSR: 500px wide) -> Laplacian variance; blank pages (mean gray > 245) skipped; threshold configurable (SATP/TSSR: 300, default: 1000); returns list of blurry page numbers with variance value"],
    ["_rule_min_photos_per_category", "sharepoint_manager.py:582",
     "Count photos per section heading group",
     "Groups pages by \\d+\\.\\d+ heading or keyword-matched title; counts embedded images >= min_dim (200x200px) per group; page 1 excluded; overflow pages (0 photos, no heading, < 50 chars text) skipped; each group must have >= min_photos (default 1)"],
    ["_rule_has_signature", "sharepoint_manager.py:689",
     "Find signature images or text in document; exclude N/A placeholders",
     "Param: start_page=1 (1-based), end_page=last, last_page_only=False. Pass 1: find 'Signature' labels -> check nearby images (w=30-200, h=15-100, aspect>1.2) + exclude if N/A text nearby; Pass 2: heuristic - small images in bottom 60% of page (text-based only); OCR fallback: pytesseract image_to_data with proximity check - if every 'Signature' word has an 'N/A' nearby (right, +/-60px vertical), page is excluded"],
]

for r, row in enumerate(rules_def, 3):
    for c, v in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 4, specific_fill)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 72

# ═══════════════════════════════════════════
# Sheet 3: Matrix (front)
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Matrix", 0)
check_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

doc_types = ["PO", "DN", "BL", "PL", "FT", "CC", "MOP", "SMR", "TSSR", "SATP", "PAF", "FAF"]
true, false = True, False

# ── Content rules ──
# Each row: (rule_label, [bool per doc_type])
rule_rows = [
    # Default rules
    ("1. Valid PDF",            [true]*12),
    ("2. Has Text",             [true]*12),
    ("3. File Naming",          [true]*12),
    # Content rules (per doc type)
    ("4. TSSR Content",         [false]*7 + [false, true, false, false, false]),
    ("5. SATP Content",         [false]*9 + [true, false, false]),
    ("6. PO Content",           [true] + [false]*11),
    ("7. DN Content",           [false, true] + [false]*10),
    ("8. BL Content",           [false, false, true] + [false]*9),
    ("9. PL Content",           [false]*3 + [true] + [false]*8),
    ("10. MOP Content",         [false]*6 + [true] + [false]*5),
    ("11. SMR Content",         [false]*7 + [true] + [false]*4),
    ("12. CC Content",          [false]*5 + [true] + [false]*6),
    ("13. FT Content",          [false]*4 + [true] + [false]*7),
    ("14. PAF Content",         [false]*10 + [true, false]),
    ("15. FAF Content",         [false]*11 + [true]),
    # Blur / Photo / Signature rules
    ("16. No Blurry Pages",     [false]*4 + [true] + [false]*2 + [false, true, true, false, false]),
    ("17. Min Photos/Cat.",     [false]*8 + [true] + [false]*3),
    ("18. Has Signature (full)",    [false]*6 + [true] + [false] + [true, true, false, false]),
    ("19. Has Signature (last pg)", [true, true, false, true, false, true] + [false]*4 + [true, true]),
    ("20. Has Signature (p2 only)", [false]*7 + [true] + [false]*4),
    ("21. Has Signature (p6+)",     [false]*4 + [true] + [false]*7),
]

ncols = len(doc_types) + 1
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
ws3.cell(row=1, column=1, value="Validation Rules x Document Type Matrix").font = Font(bold=True, size=13, color="FFFFFF")
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=1).alignment = header_align

ws3.cell(row=2, column=1, value="Rule")
for j, dt in enumerate(doc_types):
    ws3.cell(row=2, column=j+2, value=dt)
style_header(ws3, 2, ncols)

for i, (rule_name, checks) in enumerate(rule_rows):
    r = i + 3
    c1 = ws3.cell(row=r, column=1, value=rule_name)
    c1.font = Font(bold=True)
    c1.border = thin_border
    c1.alignment = wrap
    # Section coloring
    if i == 0:
        pass  # first row already styled
    elif i == 3:
        pass  # content section starts

    for j, chk in enumerate(checks):
        cell = ws3.cell(row=r, column=j+2, value=chr(0x2713) if chk else "")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if chk:
            cell.fill = check_fill
            cell.font = Font(bold=True, color="006100")

ws3.column_dimensions["A"].width = 28
for j in range(len(doc_types)):
    col = get_column_letter(j + 2)
    ws3.column_dimensions[col].width = 10

out = Path("validation_rules_matrix_v5.xlsx")
wb.save(str(out))
print(f"Exported to {out.resolve()}")
